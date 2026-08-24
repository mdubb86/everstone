"""Tests for doc_table.py — the .csv/.xlsx -> DuckDB converter.

Task 1's tests exercise the SERIALIZER directly rather than through DuckDB on
purpose: the sniffer's behaviour is a downstream consequence of the exact
field counts written here, so a failure at this level names the actual cause
("row 5 was written 4 fields wide") instead of the symptom DuckDB reports much
later and much more confusingly ("the whole file is one VARCHAR column").
"""
import datetime

import duckdb
import openpyxl
import pytest
from openpyxl import Workbook

from es.capabilities import doc_table


def _sheet(rows, title="Data", merges=()):
    """Build a real .xlsx in memory and hand back a READ-ONLY worksheet — the
    same mode doc_table uses in production. Round-tripping through an actual
    file matters: read-only mode's row padding comes from the workbook's
    declared dimension, which only exists once openpyxl has written and
    reparsed the sheet. A hand-built in-memory worksheet would pad
    differently and quietly test something else."""
    import io
    wb = Workbook()
    ws = wb.active
    ws.title = title
    for r in rows:
        ws.append(r)
    for m in merges:
        ws.merge_cells(m)
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    wb2 = openpyxl.load_workbook(buf, read_only=True, data_only=True)
    return wb2[title]


def _xlsx(tmp_path, sheets, name="book.xlsx"):
    """`sheets` is an ordered {title: [row, ...]} mapping."""
    wb = Workbook()
    wb.remove(wb.active)
    for title, rows in sheets.items():
        ws = wb.create_sheet(title=title)
        for r in rows:
            ws.append(r)
    p = tmp_path / name
    wb.save(p)
    return p


def _csv_file(tmp_path, text, name="data.csv"):
    p = tmp_path / name
    p.write_text(text, encoding="utf-8")
    return p


def _by_sheet(result):
    return {t["sheet"]: t for t in result["tables"]}


def _lines(ws, tmp_path, name="out.csv"):
    p = tmp_path / name
    doc_table.sheet_to_csv(ws, p)
    return p.read_text(encoding="utf-8").splitlines()


def test_a_banner_row_is_written_as_one_field(tmp_path):
    ws = _sheet([
        ["Quarterly Revenue Report"],
        ["id", "name", "amount"],
        [1, "alice", 10.5],
    ])
    lines = _lines(ws, tmp_path)
    # One field, not "Quarterly Revenue Report,," — the column-count contrast
    # against the 3-field rows below is the ONLY signal DuckDB's sniffer has
    # that this line is a preamble and not data.
    assert lines[0] == "Quarterly Revenue Report"
    assert lines[1] == "id,name,amount"


def test_a_data_row_with_an_empty_last_cell_keeps_full_width(tmp_path):
    """The regression guard for the bug that collapsed an entire file into a
    single column: trimming trailing empties from EVERY row made this row 2
    fields where its neighbours had 3, and the sniffer then read the whole
    file as one VARCHAR column — silently, with no error anywhere."""
    ws = _sheet([
        ["id", "name", "note"],
        [1, "alice", "first"],
        [2, "bob", None],
        [3, "carol", "third"],
    ])
    lines = _lines(ws, tmp_path)
    assert lines[2] == "2,bob,"
    assert all(len(ln.split(",")) == 3 for ln in lines)


def test_a_blank_row_is_a_blank_line(tmp_path):
    ws = _sheet([
        ["Report"],
        [],
        ["id", "name"],
        [1, "alice"],
    ])
    lines = _lines(ws, tmp_path)
    assert lines[1] == ""


def test_a_merged_cell_banner_behaves_like_a_banner(tmp_path):
    """A real spreadsheet's title is usually merged across the data columns.
    openpyxl reports the merge's value in the top-left cell and None in the
    rest, so it arrives here as an ordinary one-non-empty-cell row — this
    pins that it is not accidentally padded to full width by the merge."""
    ws = _sheet([
        ["Quarterly Revenue Report", None, None],
        ["id", "name", "amount"],
        [1, "alice", 10.5],
    ], merges=("A1:C1",))
    lines = _lines(ws, tmp_path)
    assert lines[0] == "Quarterly Revenue Report"


def test_a_single_column_sheet_is_not_mangled(tmp_path):
    """The width > 1 guard. In a genuinely one-column sheet EVERY row has
    exactly one non-empty cell — treating each as a banner would be
    harmless-looking and completely wrong, since there is no wider data row
    for them to contrast against."""
    ws = _sheet([["name"], ["alice"], ["bob"]])
    lines = _lines(ws, tmp_path)
    assert lines == ["name", "alice", "bob"]


def test_a_subtotal_row_inside_the_data_is_padded_not_treated_as_a_banner(tmp_path):
    """A row where only one column is filled is a BANNER at the top of a
    sheet and a SUBTOTAL in the middle of one — and the difference is not
    cosmetic. Written as one field mid-data it collapses the whole file:
    `id,note / 1 / 2,here` came back from DuckDB as a single VARCHAR column
    named `id,note`, holding ('1',) and ('2,here',). Measured, not
    hypothesized."""
    ws = _sheet([
        ["Quarterly Revenue Report"],
        ["id", "name", "amount"],
        [1, "alice", 10.5],
        [None, None, 10.5],          # a subtotal: only the amount column
        [2, "bob", 20.0],
    ])
    lines = _lines(ws, tmp_path)
    assert lines[0] == "Quarterly Revenue Report", "the real banner still is one"
    assert lines[3] == ",,10.5"
    assert all(len(ln.split(",")) == 3 for ln in lines[1:])


def test_dates_serialize_as_iso_and_none_as_empty(tmp_path):
    ws = _sheet([
        ["id", "when", "note"],
        [1, datetime.datetime(2026, 1, 2), None],
        [2, datetime.datetime(2026, 1, 3, 14, 30), "later"],
    ])
    lines = _lines(ws, tmp_path)
    # A date-formatted cell comes back from openpyxl as a datetime at
    # midnight — writing it as a bare date keeps DuckDB's sniffer inferring
    # DATE rather than TIMESTAMP for what the spreadsheet showed as a date.
    assert lines[1] == "1,2026-01-02,"
    assert lines[2] == "2,2026-01-03 14:30:00,later"


def test_a_value_containing_the_delimiter_is_quoted(tmp_path):
    ws = _sheet([
        ["id", "note"],
        [1, "hello, world"],
        [2, 'she said "hi"'],
    ])
    lines = _lines(ws, tmp_path)
    assert lines[1] == '1,"hello, world"'
    assert lines[2] == '2,"she said ""hi"""'


def test_an_empty_sheet_writes_nothing(tmp_path):
    ws = _sheet([])
    p = tmp_path / "empty.csv"
    doc_table.sheet_to_csv(ws, p)
    assert p.read_text(encoding="utf-8") == ""


# ---------------------------------------------------------------- build()


def test_a_csv_preamble_is_skipped_and_types_inferred(tmp_path):
    src = _csv_file(tmp_path, (
        "Quarterly Revenue Report\n"
        "generated 2026-08-24 by acme\n"
        "\n"
        "id,name,amount,when\n"
        "1,alice,10.5,2026-01-02\n"
        "2,bob,20.25,2026-01-03\n"
    ))
    adir = tmp_path / "art"
    adir.mkdir()
    result = doc_table.build(src, adir)

    assert (adir / doc_table.DB_NAME).is_file()
    t = result["tables"][0]
    assert [c["name"] for c in t["columns"]] == ["id", "name", "amount", "when"]
    assert [c["type"] for c in t["columns"]] == ["BIGINT", "VARCHAR", "DOUBLE", "DATE"]
    assert t["rows"] == 2
    # 1-based line number of the header row itself — three lines of preamble
    # precede it, so it is line 4. Stated as a LINE NUMBER (not a skip count)
    # so it can be cross-checked directly against the `cells` table, which is
    # also 1-based.
    assert t["header_row"] == 4


def test_a_headerless_csv_reports_no_header_row(tmp_path):
    src = _csv_file(tmp_path, "1,2,3\n4,5,6\n")
    adir = tmp_path / "art"
    adir.mkdir()
    t = doc_table.build(src, adir)["tables"][0]
    assert t["header_row"] is None
    assert [c["name"] for c in t["columns"]] == ["column0", "column1", "column2"]


def test_each_sheet_becomes_its_own_table_with_its_own_schema(tmp_path):
    src = _xlsx(tmp_path, {
        "Sales": [["id", "amount"], [1, 10.5], [2, 20.5]],
        "People": [["name", "city"], ["alice", "austin"]],
    })
    adir = tmp_path / "art"
    adir.mkdir()
    tables = _by_sheet(doc_table.build(src, adir))

    assert set(tables) == {"Sales", "People"}
    assert [c["type"] for c in tables["Sales"]["columns"]] == ["BIGINT", "DOUBLE"]
    assert [c["name"] for c in tables["People"]["columns"]] == ["name", "city"]
    assert tables["Sales"]["rows"] == 2
    assert tables["People"]["rows"] == 1


@pytest.mark.parametrize("title", [
    "Q1 Sales",           # spaces
    "2024 Budget",        # leading digit
    "order",              # a reserved SQL keyword
    "P&L (draft)",        # punctuation
    "Ventas Año",         # non-ASCII
])
def test_a_sheet_name_becomes_a_usable_table_name(tmp_path, title):
    """Whatever the mapping is, the name it produces must actually work as an
    unquoted identifier in the SQL the agent will write — that is the only
    property that matters here, so assert it by RUNNING a query rather than
    by pinning a particular slug."""
    src = _xlsx(tmp_path, {title: [["id"], [1]]})
    adir = tmp_path / "art"
    adir.mkdir()
    t = doc_table.build(src, adir)["tables"][0]

    assert t["sheet"] == title
    con = duckdb.connect(str(adir / doc_table.DB_NAME), read_only=True)
    try:
        assert con.execute(f"SELECT count(*) FROM {t['table']}").fetchone()[0] == 1
    finally:
        con.close()


def test_the_sheet_to_table_mapping_is_returned_not_inferred(tmp_path):
    """An identifier the agent reconstructs is one it gets wrong. Two sheets
    that slugify to the same name must still be distinguishable, and the
    receipt has to say which is which."""
    src = _xlsx(tmp_path, {
        "Q1 Sales": [["id"], [1]],
        "Q1-Sales": [["id"], [2], [3]],
    })
    adir = tmp_path / "art"
    adir.mkdir()
    tables = _by_sheet(doc_table.build(src, adir))

    names = [t["table"] for t in tables.values()]
    assert len(set(names)) == 2, "deduplication must not collapse two sheets"
    con = duckdb.connect(str(adir / doc_table.DB_NAME), read_only=True)
    try:
        for sheet, expect in (("Q1 Sales", 1), ("Q1-Sales", 2)):
            got = con.execute(
                f"SELECT count(*) FROM {tables[sheet]['table']}").fetchone()[0]
            assert got == expect, f"{sheet} maps to the wrong table"
    finally:
        con.close()


def test_an_empty_sheet_is_a_table_with_zero_rows_not_an_error(tmp_path):
    src = _xlsx(tmp_path, {"Blank": [], "Real": [["id"], [1]]})
    adir = tmp_path / "art"
    adir.mkdir()
    tables = _by_sheet(doc_table.build(src, adir))

    assert tables["Blank"]["rows"] == 0
    con = duckdb.connect(str(adir / doc_table.DB_NAME), read_only=True)
    try:
        assert con.execute(
            f"SELECT count(*) FROM {tables['Blank']['table']}").fetchone()[0] == 0
    finally:
        con.close()


def test_a_large_sheet_converts_in_full_with_no_row_cap(tmp_path):
    """The bug that started this: 40,000 rows of spreadsheet became a capped
    Markdown table, and the agent confidently answered a count question with
    the size of the cap."""
    from openpyxl import Workbook as WB
    wb = WB(write_only=True)
    ws = wb.create_sheet(title="Txns")
    ws.append(["id", "amount"])
    for i in range(1, 40001):
        ws.append([i, float(i)])
    src = tmp_path / "big.xlsx"
    wb.save(src)

    adir = tmp_path / "art"
    adir.mkdir()
    t = doc_table.build(src, adir)["tables"][0]
    assert t["rows"] == 40000

    con = duckdb.connect(str(adir / doc_table.DB_NAME), read_only=True)
    try:
        assert con.execute(f"SELECT count(*) FROM {t['table']}").fetchone()[0] == 40000
        assert con.execute(
            f"SELECT sum(amount) FROM {t['table']}").fetchone()[0] == 800020000.0
    finally:
        con.close()


def test_rebuilding_over_an_existing_database_replaces_it(tmp_path):
    """The artifact directory is keyed by a content hash, so a rebuild always
    means the same content — but a half-written database left by a crashed
    conversion must not be appended to or reused."""
    adir = tmp_path / "art"
    adir.mkdir()
    src = _xlsx(tmp_path, {"Sales": [["id"], [1]]})
    doc_table.build(src, adir)
    t = doc_table.build(src, adir)["tables"][0]

    con = duckdb.connect(str(adir / doc_table.DB_NAME), read_only=True)
    try:
        assert con.execute(f"SELECT count(*) FROM {t['table']}").fetchone()[0] == 1
    finally:
        con.close()


# ----------------------------------------------------------------- cells


def _cells(adir):
    con = duckdb.connect(str(adir / doc_table.DB_NAME), read_only=True)
    try:
        return con.execute(
            f"SELECT sheet, row, col, ref, value FROM {doc_table.CELLS_TABLE} "
            "ORDER BY sheet, row, col").fetchall()
    finally:
        con.close()


def test_cells_recovers_the_real_header_when_detection_gets_it_wrong(tmp_path):
    """The point of the whole table, so it is tested first.

    Without an escape hatch a misdetected header is a silently wrong answer to
    every later question. This file misdetects for real (verified against
    DuckDB 1.5.5): a subtitle line with the same field count as the data wins
    the header vote, and the ACTUAL header becomes row one of the data.
    """
    src = _csv_file(tmp_path, (
        "Region,Q1,Q2\n"
        "id,name,amount\n"
        "1,alice,10\n"
        "2,bob,20\n"
    ))
    adir = tmp_path / "art"
    adir.mkdir()
    t = doc_table.build(src, adir)["tables"][0]

    # The misdetection itself — pinned so the test still means something if
    # DuckDB's sniffer improves and this file stops fooling it.
    assert [c["name"] for c in t["columns"]] == ["Region", "Q1", "Q2"]
    assert t["header_row"] == 1

    # ...and the recovery: the real header is right there, addressable by row.
    con = duckdb.connect(str(adir / doc_table.DB_NAME), read_only=True)
    try:
        row2 = con.execute(
            f"SELECT value FROM {doc_table.CELLS_TABLE} "
            "WHERE row = 2 ORDER BY col").fetchall()
    finally:
        con.close()
    assert [v[0] for v in row2] == ["id", "name", "amount"]


def test_every_non_empty_cell_is_addressable_by_sheet_row_and_column(tmp_path):
    src = _xlsx(tmp_path, {
        "Sales": [["id", "amount"], [1, 10.5]],
        "Notes": [["only"]],
    })
    adir = tmp_path / "art"
    adir.mkdir()
    doc_table.build(src, adir)

    assert _cells(adir) == [
        ("Notes", 1, 1, "A1", "only"),
        ("Sales", 1, 1, "A1", "id"),
        ("Sales", 1, 2, "B1", "amount"),
        ("Sales", 2, 1, "A2", "1"),
        ("Sales", 2, 2, "B2", "10.5"),
    ]


def test_an_empty_cell_is_absent_rather_than_stored(tmp_path):
    """Sparse on purpose: a 40,000-row sheet of 20 columns would otherwise
    put 800,000 rows in this table, most of them recording nothing. An
    absent (row, col) means empty — which is also what a spreadsheet means
    by an empty cell."""
    src = _xlsx(tmp_path, {"Sales": [["id", "note"], [1, None], [2, "here"]]})
    adir = tmp_path / "art"
    adir.mkdir()
    doc_table.build(src, adir)

    refs = [c[3] for c in _cells(adir)]
    assert "B2" not in refs
    assert refs == ["A1", "B1", "A2", "A3", "B3"]


@pytest.mark.parametrize("col,ref", [
    (1, "A"), (26, "Z"), (27, "AA"), (28, "AB"), (52, "AZ"),
    (53, "BA"), (702, "ZZ"), (703, "AAA"),
])
def test_a1_notation_matches_what_a_spreadsheet_shows(col, ref):
    assert doc_table.a1_ref(col, 7) == f"{ref}7"


def test_cell_values_are_raw_text_not_typed(tmp_path):
    """`cells` is the check ON type inference, so it must not be subject to
    it: everything is VARCHAR, exactly the text the typed table was built
    from. A number that lost precision, a date read as a string, a column
    that came out entirely NULL — all of them are diagnosable only if this
    table shows what was actually there."""
    src = _csv_file(tmp_path, "id,when\n007,2026-01-02\n")
    adir = tmp_path / "art"
    adir.mkdir()
    doc_table.build(src, adir)

    con = duckdb.connect(str(adir / doc_table.DB_NAME), read_only=True)
    try:
        types = [r[1] for r in con.execute(
            f"DESCRIBE {doc_table.CELLS_TABLE}").fetchall()]
        assert types[-1] == "VARCHAR"
        # The typed table read "007" as the number 7; cells kept the zeros.
        assert con.execute(
            f"SELECT value FROM {doc_table.CELLS_TABLE} "
            "WHERE ref = 'A2'").fetchone()[0] == "007"
    finally:
        con.close()


def test_cell_rows_line_up_with_the_reported_header_row(tmp_path):
    """The cross-check has to actually work: `header_row` is only useful if
    looking that row up in `cells` shows the header."""
    src = _csv_file(tmp_path, (
        "Quarterly Revenue Report\n"
        "\n"
        "id,name\n"
        "1,alice\n"
    ))
    adir = tmp_path / "art"
    adir.mkdir()
    t = doc_table.build(src, adir)["tables"][0]

    con = duckdb.connect(str(adir / doc_table.DB_NAME), read_only=True)
    try:
        header = con.execute(
            f"SELECT value FROM {doc_table.CELLS_TABLE} WHERE row = ? "
            "ORDER BY col", [t["header_row"]]).fetchall()
    finally:
        con.close()
    assert [v[0] for v in header] == ["id", "name"]


def test_a_sheet_with_a_subtotal_row_still_builds_a_real_table(tmp_path):
    """The serializer test above pins the field count; this pins what DuckDB
    then does with it. Without the preamble guard the whole sheet came back
    as ONE VARCHAR column, and every later query against it would have been
    confidently wrong."""
    src = _xlsx(tmp_path, {"Sales": [
        ["Quarterly Revenue Report"],
        ["id", "name", "amount"],
        [1, "alice", 10.5],
        [None, None, 10.5],
        [2, "bob", 20.0],
    ]})
    adir = tmp_path / "art"
    adir.mkdir()
    t = doc_table.build(src, adir)["tables"][0]

    assert [c["name"] for c in t["columns"]] == ["id", "name", "amount"]
    assert t["rows"] == 3
    con = duckdb.connect(str(adir / doc_table.DB_NAME), read_only=True)
    try:
        assert con.execute(
            f"SELECT sum(amount) FROM {t['table']}").fetchone()[0] == 41.0
    finally:
        con.close()


# ----------------------------------------------------------------- query()


@pytest.fixture
def txns(tmp_path):
    """A two-sheet workbook built once for the query tests: 200 transactions
    and the customers they belong to."""
    rows = [["id", "customer_id", "amount"]]
    for i in range(1, 201):
        rows.append([i, (i % 3) + 1, float(i)])
    src = _xlsx(tmp_path, {
        "Txns": rows,
        "Customers": [["id", "name"], [1, "alice"], [2, "bob"], [3, "carol"]],
    })
    adir = tmp_path / "art"
    adir.mkdir()
    doc_table.build(src, adir)
    return adir


def test_an_aggregate_answers_in_one_row(txns):
    """The motivating case. Previously this meant paging a capped Markdown
    table and adding numbers by hand; now it is one call returning one row."""
    out = doc_table.query(txns, "SELECT count(*), sum(amount) FROM txns WHERE amount > 100")
    assert out["rows"] == [[100, 15050.0]]
    assert out["truncated"] is False


def test_a_join_across_two_sheets_works(txns):
    out = doc_table.query(txns, (
        "SELECT c.name, count(*) AS n FROM txns t "
        "JOIN customers c ON c.id = t.customer_id GROUP BY c.name ORDER BY c.name"))
    assert [r[0] for r in out["rows"]] == ["alice", "bob", "carol"]
    assert sum(r[1] for r in out["rows"]) == 200


@pytest.mark.parametrize("sql", [
    "INSERT INTO txns VALUES (999, 1, 1.0)",
    "UPDATE txns SET amount = 0",
    "DELETE FROM txns",
    "DROP TABLE txns",
    "CREATE TABLE evil AS SELECT 1",
    "ATTACH '/tmp/other.duckdb' AS o",
    "COPY (SELECT 1) TO '/tmp/exfil.csv'",
    "SELECT 1; DROP TABLE txns",
])
def test_anything_that_is_not_a_read_is_refused(txns, sql):
    with pytest.raises(doc_table.NotAQuery) as e:
        doc_table.query(txns, sql)
    assert "read" in str(e.value).lower()


def test_the_database_is_unchanged_after_a_refused_write(txns):
    """Refusing is only meaningful if nothing happened. Checked separately
    from the refusal itself so a change that started executing before
    raising would still be caught."""
    for sql in ("DELETE FROM txns", "DROP TABLE txns"):
        with pytest.raises(doc_table.NotAQuery):
            doc_table.query(txns, sql)
    assert doc_table.query(txns, "SELECT count(*) FROM txns")["rows"] == [[200]]


def test_reading_a_file_off_the_host_is_refused(txns):
    """The statement allowlist alone would let this through — it is a SELECT.
    External filesystem access is disabled on the connection itself, which is
    what actually stops it."""
    with pytest.raises(doc_table.QueryFailed) as e:
        doc_table.query(txns, "SELECT * FROM read_csv('/etc/hostname')")
    assert "disabled" in str(e.value).lower() or "permission" in str(e.value).lower()


def test_a_query_with_no_limit_gets_one_and_says_so(txns):
    # 600 rows from a fixture of 200, so the cap bites regardless of what the
    # fixture's own size happens to be — an earlier version of this test used
    # `SELECT * FROM txns`, which returned exactly MAX_QUERY_ROWS and could
    # therefore never have observed a missing cap.
    out = doc_table.query(txns, "SELECT t.id, c.name FROM txns t, customers c")
    assert len(out["rows"]) == doc_table.MAX_QUERY_ROWS
    assert out["truncated"] is True
    assert out["row_count"] == doc_table.MAX_QUERY_ROWS


def test_the_agents_own_limit_is_respected(txns):
    out = doc_table.query(txns, "SELECT * FROM txns ORDER BY id LIMIT 5")
    assert len(out["rows"]) == 5
    assert out["truncated"] is False
    assert [r[0] for r in out["rows"]] == [1, 2, 3, 4, 5]


def test_an_unknown_table_error_names_the_real_tables(txns):
    with pytest.raises(doc_table.QueryFailed) as e:
        doc_table.query(txns, "SELECT * FROM transactions")
    msg = str(e.value)
    assert "txns" in msg and "customers" in msg
    # The mapping too — the agent asked for the SHEET name, so the error has
    # to bridge from what it typed to what the table is called.
    assert "Txns" in msg


def test_introspection_commands_are_allowed(txns):
    """DESCRIBE/SHOW/SUMMARIZE all classify as SELECT to DuckDB, so the agent
    gets them for free — worth pinning, since they are how it recovers a
    schema it no longer has the receipt for."""
    assert doc_table.query(txns, "DESCRIBE txns")["rows"]
    names = [r[0] for r in doc_table.query(txns, "SHOW TABLES")["rows"]]
    assert "txns" in names and doc_table.META_TABLE in names


def test_a_runaway_query_is_stopped(txns):
    """A three-way cross join is 8e6 rows here and unbounded in general.
    DuckDB has no statement_timeout setting; the bound comes from calling
    interrupt() on the connection from a timer, which was measured to stop a
    genuine runaway in exactly the time allowed."""
    # The predicate matters: a bare `count(*)` over a cross join is answered
    # from cardinality alone without materializing anything, so it finishes
    # instantly and proves nothing. This forces 1.6e9 comparisons.
    with pytest.raises(doc_table.QueryTimeout):
        doc_table.query(txns, (
            "SELECT count(*) FROM txns a, txns b, txns c, txns d "
            "WHERE a.amount * b.amount > c.amount * d.amount"), timeout=0.25)


def test_values_come_back_json_safe(txns, tmp_path):
    """Whatever this returns is going into an MCP envelope, so a date or a
    decimal has to already be a string or a number — not a Python object that
    json.dumps refuses at the very end of the call."""
    import json
    src = _csv_file(tmp_path, "id,when\n1,2026-01-02\n", name="d.csv")
    adir = tmp_path / "art2"
    adir.mkdir()
    doc_table.build(src, adir)
    out = doc_table.query(adir, "SELECT * FROM d")
    assert out["rows"] == [[1, "2026-01-02"]]
    json.dumps(out)
