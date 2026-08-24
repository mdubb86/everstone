"""`.csv`/`.xlsx` -> a queryable DuckDB database.

These two formats stop being documents here. A spreadsheet converted to a
Markdown pipe table is a document the agent has to READ — 40,000 rows of it,
one window at a time, to answer "how many transactions over $500 in
September". As a table it is one `SELECT count(*) ... WHERE` returning one
row. The agent writes competent SQL out of the box, so the useful thing to
hand it is a database, not prose about a database.

Both formats reach DuckDB through the SAME door: its CSV sniffer. `.xlsx` is
read by openpyxl and serialized to CSV first rather than going through
DuckDB's own `read_xlsx`, which was measured and rejected — given a sheet
with an ordinary title banner it returned ONE column, named after the banner,
containing ZERO rows. The sniffer, by contrast, needed no configuration at
all for the same shape: it reported `SkipRows: 4` and inferred
`BIGINT/VARCHAR/DOUBLE/DATE` correctly. One code path for both formats also
means header detection, type inference and preamble handling are the same
machinery, tested once.

**The serialization rule below is load-bearing — do not "tidy" it.** The
sniffer identifies a preamble purely by COLUMN-COUNT CONTRAST: a line with
fewer fields than the block beneath it is not data. So:

    blank row                                     -> a blank line
    one non-empty cell, BEFORE the first data row -> ONE field (a banner)
    anything else                                 -> PADDED to full width

An earlier version trimmed trailing empty cells from every row. That made a
data row whose last column happened to be empty 4 fields wide where its
neighbours were 5 — and DuckDB then read the entire file as a SINGLE VARCHAR
column. Nothing raised, nothing warned; the data was simply wrong.

The "before the first data row" qualifier is the same failure, found the same
way. Written without it, a SUBTOTAL row — one where only the amount column is
filled, which is ordinary in a real spreadsheet — collapsed to one field in
the middle of the data, and `id,note / 1 / 2,here` came back as a single
column named `id,note` holding `('1',)` and `('2,here',)`. A banner is a
top-of-sheet thing; once a full-width row has been written the block has
started, and a narrow row inside it is data with empty cells, not a title.
"""
import csv
import datetime
import tempfile
import threading
from pathlib import Path
from typing import List, Optional, Tuple
from xml.etree import ElementTree
from zipfile import BadZipFile

import duckdb
import openpyxl
from openpyxl.utils.exceptions import InvalidFileException

from es.capabilities import doc_support

# The stdlib csv module's default field_size_limit (128 KiB) is a parser
# safety valve, not a real document-shape limit — a single free-text CSV
# column (a pasted comment, a description field) can legitimately exceed it,
# and hitting the default raises `_csv.Error: field larger than field limit`.
# This module is the only remaining place es parses a CSV with Python rather
# than DuckDB (_write_cells, which re-reads the exact bytes DuckDB was fed),
# so the process-wide setting lives here now — it used to sit in docs.py,
# next to a doc_text CSV converter that no longer exists. 10 MiB comfortably
# covers any realistic single field while staying a real ceiling: a CSV with
# an unterminated quote makes csv.reader treat everything from that quote to
# EOF as ONE field, so it must hit SOME limit rather than buffer the rest of
# a 50MB upload into one Python string.
CSV_FIELD_SIZE_LIMIT = 10 * 1024 * 1024
csv.field_size_limit(CSV_FIELD_SIZE_LIMIT)


def _cell_text(value) -> str:
    """One spreadsheet cell as the text CSV should carry.

    Dates are the case worth explaining. openpyxl hands back a
    `datetime.datetime` for BOTH a date-formatted cell and a real timestamp —
    Excel stores one numeric type for both and the distinction lives only in
    the number format. A date-formatted cell therefore arrives as midnight,
    and writing it out as `2026-01-02 00:00:00` makes DuckDB infer TIMESTAMP
    for a column the spreadsheet displayed as dates. Emitting the bare date
    when the time component is exactly midnight recovers the DATE the user
    actually sees. The cost is a genuine timestamp that happens to land on
    midnight losing a `00:00:00` it can be read back as anyway — DuckDB parses
    a bare date into a TIMESTAMP column without complaint, so a column mixing
    both still comes out right.
    """
    if value is None:
        return ""
    if isinstance(value, bool):
        # Before the int check: bool IS an int in Python, and "True" is not a
        # value DuckDB reads back as BOOLEAN as reliably as "true".
        return "true" if value else "false"
    if isinstance(value, datetime.datetime):
        if (value.hour, value.minute, value.second, value.microsecond) == (0, 0, 0, 0):
            return value.date().isoformat()
        return value.isoformat(sep=" ")
    if isinstance(value, (datetime.date, datetime.time)):
        return value.isoformat()
    return str(value)


def _rows(ws, source: Optional[Path]):
    """Rows of `ws`, with lazy XML parse errors translated when we know which
    file to blame. `source=None` is the direct-serialization path (tests, and
    any caller holding a worksheet with no file behind it) — there is nothing
    useful to name in an error there, so the raw iterator is fine."""
    if source is None:
        return ws.iter_rows(values_only=True)
    return _safe_rows(ws, source)


def _sheet_width(ws, source: Optional[Path] = None) -> int:
    """The full data width every non-banner row is padded to.

    `ws.max_column` comes from the workbook's DECLARED dimension, and in
    read-only mode that same declaration is what makes openpyxl pad each row
    it yields to a uniform length — so the two agree by construction and the
    padding below is normally a no-op safety net. It stays anyway for the
    workbook whose declared dimension is too SMALL: a row longer than `width`
    is written in full (the negative padding count degrades to no padding
    rather than truncating), because losing a column of real data would be
    far worse than the ragged row the sniffer might then misread.

    Do NOT call `ws.reset_dimensions()` to "fix" a suspect declaration: it
    makes iter_rows yield RAGGED rows (measured: a 5-column sheet gave rows
    of 1, 0, 5, 5, 4, 5), which is exactly the trailing-empty asymmetry the
    module docstring's rule exists to prevent.

    A workbook written by `openpyxl.Workbook(write_only=True)` has NO
    declared dimension at all — `max_column` is None — and its rows then
    arrive at their natural, ragged lengths (measured: 1, 3, 3, 2 for a
    sheet whose last row had an empty final cell). Falling back to a width of
    1 there would write those rows out ragged and collapse the whole file to
    one VARCHAR column, so this scans the sheet once to find the real width
    instead. Read-only worksheets re-parse their XML on each `iter_rows`, so
    the caller's own pass still sees every row; the cost is reading the sheet
    twice, and it is paid ONLY by dimension-less workbooks.
    """
    width = ws.max_column
    if isinstance(width, int) and width > 0:
        return width
    return max((len(row) for row in _rows(ws, source)), default=1) or 1


def sheet_to_csv(ws, path: Path, source: Optional[Path] = None) -> int:
    """Serialize one worksheet to `path` per the module docstring's rule.
    Returns the number of rows written.

    Reads with `data_only=True`, which returns Excel's CACHED formula results
    — so a workbook written by a tool that never opened Excel (a report
    exported by a web app, most commonly) has no cache and its formula cells
    arrive EMPTY. There is no fix at this layer: the formula's result does
    not exist in the file. The `cells` table is the backstop — it records
    what each cell actually held, so a column of blanks is at least
    diagnosable rather than silently zero.

    Trailing blank rows are dropped rather than written. An .xlsx routinely
    declares a dimension larger than its real data (openpyxl then pads out
    the difference), and hundreds of empty lines at the end of the file are
    noise the sniffer has to reason past. Blank rows BETWEEN data are kept —
    those are structure, often the very gap that separates a preamble from
    its table.
    """
    width = _sheet_width(ws, source)
    written = 0
    pending_blanks = 0
    # Flipped by the first full-width row. After that a one-cell row is a
    # subtotal or a sparse data row, never a banner — see the module
    # docstring for the file-collapsing bug that distinction prevents.
    in_preamble = True
    with open(path, "w", newline="", encoding="utf-8") as fh:
        # "\n" rather than the dialect default "\r\n" so a blank row is one
        # byte and every line ends the same way, whichever branch wrote it.
        writer = csv.writer(fh, lineterminator="\n")
        for row in _rows(ws, source):
            cells = [_cell_text(v) for v in row]
            non_empty = [c for c in cells if c != ""]
            if not non_empty:
                # Held back, not written: if nothing follows, these are the
                # sheet's trailing padding and belong nowhere.
                pending_blanks += 1
                continue
            for _ in range(pending_blanks):
                writer.writerow([])
                written += 1
            pending_blanks = 0
            if in_preamble and len(non_empty) == 1 and width > 1:
                writer.writerow([non_empty[0]])
            else:
                writer.writerow(cells + [""] * (width - len(cells)))
                in_preamble = False
            written += 1
    return written


DB_NAME = "data.duckdb"

# Table names this module owns. Seeded into the used-name set before any
# sheet is slugified, so a sheet actually called "Cells" cannot take the name
# and leave the metadata table to be renamed out from under the agent — the
# tables it is TOLD about in the receipt must be the ones it can query.
META_TABLE = "tables_meta"
CELLS_TABLE = "cells"
_RESERVED_TABLE_NAMES = (META_TABLE, CELLS_TABLE)


def _reserved_keywords(con) -> set:
    """DuckDB's own reserved-word list rather than a hardcoded copy — the
    list is version-specific (75 entries in 1.5.5) and a stale local copy
    would fail exactly where it matters: silently, at query time, on a sheet
    innocently named "Order"."""
    return {r[0].lower() for r in con.execute(
        "SELECT keyword_name FROM duckdb_keywords() "
        "WHERE keyword_category = 'reserved'").fetchall()}


def _slug(name: str, used: set, reserved: set) -> str:
    """A sheet name as a table identifier the agent can type UNQUOTED.

    Unquoted is the requirement worth stating: the agent writes SQL from the
    receipt, and an identifier needing quotes is one it will sooner or later
    write without them. Unicode letters are kept rather than stripped —
    DuckDB accepts them unquoted (verified: `CREATE TABLE ventas_año` works),
    so a Spanish or Japanese sheet name stays recognizable instead of
    decaying into `ventas_a_o`.
    """
    slug = "".join(ch if ch.isalnum() else "_" for ch in name.lower())
    while "__" in slug:
        slug = slug.replace("__", "_")
    slug = slug.strip("_")
    if not slug:
        slug = "sheet"
    if slug[0].isdigit() or slug in reserved:
        slug = "t_" + slug
    candidate = slug
    n = 2
    while candidate in used:
        candidate = f"{slug}_{n}"
        n += 1
    used.add(candidate)
    return candidate


# The first chunk of a file to inspect when deciding whether it is text at
# all. A NUL byte in the first 8 KiB is git's own binary heuristic, and it is
# the right one here: no text encoding produces NUL, so its presence is a
# reliable "this is not a spreadsheet someone exported" signal — which is
# what lets the latin-1 fallback below be safe.
_BINARY_SNIFF_BYTES = 8192


def _looks_like_text(path: Path) -> bool:
    try:
        with open(path, "rb") as fh:
            return b"\x00" not in fh.read(_BINARY_SNIFF_BYTES)
    except OSError:
        return False


def _read_csv_args(csv_path: Path, encoding: Optional[str]) -> str:
    return ("read_csv(?)" if encoding is None
            else f"read_csv(?, encoding='{encoding}')")


def _pick_encoding(con, csv_path: Path) -> Optional[str]:
    """None for UTF-8 (DuckDB's default), or 'latin-1' for a file that is not
    valid UTF-8 but is still plainly text.

    Exporting a CSV from Excel on Windows produces cp1252, not UTF-8, and
    DuckDB rejects the whole file the moment it hits one accented character —
    "Invalid unicode (byte sequence mismatch)". That is a real user's real
    spreadsheet, so falling back matters. latin-1 decodes ANY byte sequence
    and therefore never fails, which is exactly why it must be gated on
    _looks_like_text: without that, a PDF misnamed `.csv` would "succeed" and
    produce a table of mojibake instead of being reported as unreadable.

    We ask by TRYING, rather than by guessing from the bytes: the sniffer's
    own verdict on the actual file is more reliable than any charset
    heuristic we would write here.
    """
    try:
        con.execute(f"SELECT * FROM sniff_csv(?)", [str(csv_path)]).fetchone()
        return None
    except duckdb.Error:
        if not _looks_like_text(csv_path):
            raise
        return "latin-1"


def _sniff(con, csv_path: Path, encoding: Optional[str] = None) -> dict:
    """DuckDB's own sniffer decisions for one CSV — the delimiter it chose,
    whether it found a header, and how many lines it skipped to get there.

    This IS `tables_meta`: the value of exposing it is that a MISPARSE
    becomes visible. Header detection on a messy real spreadsheet is a
    guess, and a wrong guess is otherwise a silently wrong answer to every
    later question. Reported alongside `cells`, the agent can cross-check
    "the header is on line 4" against what line 4 actually holds.
    """
    sniff = ("sniff_csv(?)" if encoding is None
             else f"sniff_csv(?, encoding='{encoding}')")
    cur = con.execute(f"SELECT * FROM {sniff}", [str(csv_path)])
    names = [d[0] for d in cur.description]
    row = cur.fetchone()
    if row is None:
        return {"delimiter": ",", "header_row": None}
    info = dict(zip(names, row))
    skip = info.get("SkipRows") or 0
    has_header = bool(info.get("HasHeader"))
    return {
        "delimiter": info.get("Delimiter") or ",",
        # 1-based LINE NUMBER of the header row itself, not a skip count —
        # the same numbering `cells` uses, so the two can be compared
        # directly without the agent having to know which is off by one.
        "header_row": skip + 1 if has_header else None,
    }


# A password-protected .xlsx is not a zip at all: Excel wraps the real
# package inside an OLE2 compound document. openpyxl reports that as the same
# "File is not a zip file" it gives any corrupt file, so the magic bytes are
# the only way to tell the agent the one thing it can act on.
_OLE2_MAGIC = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"

# Every exception type openpyxl raises for a file it cannot read. Kept
# explicit rather than a bare `except Exception` so a bug in OUR row handling
# can never be caught here and relabelled "corrupt file" — the whole point of
# doc_support.ParseFailed's scoping contract.
_PARSE_ERRORS = (OSError, ValueError, KeyError, AttributeError, TypeError,
                 BadZipFile, InvalidFileException, ElementTree.ParseError)


def _raise_parse_failed(source: Path, exc: Exception) -> None:
    """Always raises."""
    try:
        with open(source, "rb") as fh:
            encrypted = fh.read(len(_OLE2_MAGIC)) == _OLE2_MAGIC
    except OSError:
        encrypted = False
    if encrypted:
        raise doc_support.ParseFailed(
            f"{source.name} is password-protected — es cannot open "
            "encrypted Excel workbooks; ask the user for an unlocked copy",
            encrypted=True) from exc
    raise doc_support.ParseFailed(
        f"{source.name} could not be read as an Excel workbook — it may be "
        "corrupt, truncated, or not actually a .xlsx file; ask the user to "
        "resend it") from exc


def _open_workbook(source: Path):
    """Opens the WORKBOOK-level parts (workbook.xml, styles, shared strings)
    eagerly — but NOT a sheet's own XML, which read_only mode streams lazily.
    A truncated sheet1.xml therefore raises during ITERATION, not here; see
    _safe_rows."""
    try:
        return openpyxl.load_workbook(source, read_only=True, data_only=True)
    except _PARSE_ERRORS as e:
        _raise_parse_failed(source, e)


def _safe_rows(ws, source: Path):
    """Wrap ONLY the iterator's own `next()` — the exact point openpyxl's
    read-only reader parses the next chunk of a sheet's XML. Verified: a
    truncated sheet1.xml raises `xml.etree.ElementTree.ParseError`
    mid-iteration, never at load_workbook() time. The caller's own row
    handling stays outside this try block, so a bug there can never be
    mistaken for a corrupt file."""
    rows = ws.iter_rows(values_only=True)
    while True:
        try:
            row = next(rows)
        except StopIteration:
            return
        except _PARSE_ERRORS as e:
            _raise_parse_failed(source, e)
        yield row


def _serialize(source: Path, workdir: Path) -> List[Tuple[str, Path]]:
    """Every unit of the source as (name, csv_path), in source order.

    A `.csv` is already CSV and is handed to DuckDB UNCHANGED — re-serializing
    it would mean parsing it ourselves first, which throws away exactly the
    delimiter/quoting/embedded-newline handling DuckDB's sniffer is better at
    than we are. Its one pseudo-sheet takes the file's stem as a name, so
    `transactions.csv` reads as sheet "transactions" and a receipt has the
    same shape for both formats.
    """
    if source.suffix.lower() == ".csv":
        return [(source.stem, source)]
    wb = _open_workbook(source)
    try:
        out = []
        for ws in wb.worksheets:
            path = workdir / f"sheet{len(out):03d}.csv"
            sheet_to_csv(ws, path, source)
            out.append((ws.title, path))
        return out
    finally:
        wb.close()


def build(source: Path, adir: Path) -> dict:
    """Convert `source` into `adir/data.duckdb` and describe what was built.

    Returns `{"tables": [{sheet, table, rows, header_row, columns}, ...]}` in
    source order — `columns` is `[{name, type}, ...]` as DuckDB actually
    stored them, read back with DESCRIBE rather than echoed from the sniffer,
    so the receipt cannot claim a schema the database does not have.

    The same description is also written INTO the database as `tables_meta`,
    so the mapping survives the receipt: an agent that has the doc_id but not
    the original response can recover which table a sheet became with one
    query instead of guessing at the slug.
    """
    db_path = Path(adir) / DB_NAME
    # A crashed conversion can leave a partial database behind, and the
    # artifact dir is keyed by a content hash — so the stale file would be
    # found by the very next extract of the same document and appended to.
    if db_path.exists():
        db_path.unlink()

    with tempfile.TemporaryDirectory() as tmp:
        sheets = _serialize(Path(source), Path(tmp))
        con = duckdb.connect(str(db_path))
        try:
            reserved = _reserved_keywords(con)
            used = set(_RESERVED_TABLE_NAMES)
            tables = []
            sheets_meta = []
            for name, csv_path in sheets:
                table = _slug(name, used, reserved)
                try:
                    encoding = _pick_encoding(con, csv_path)
                    meta = _sniff(con, csv_path, encoding)
                    con.execute(
                        f'CREATE TABLE "{table}" AS SELECT * FROM '
                        + _read_csv_args(csv_path, encoding), [str(csv_path)])
                except duckdb.Error as e:
                    raise doc_support.ParseFailed(str(e)) from e
                columns = [{"name": r[0], "type": r[1]} for r in
                           con.execute(f'DESCRIBE "{table}"').fetchall()]
                rows = con.execute(
                    f'SELECT count(*) FROM "{table}"').fetchone()[0]
                tables.append({"sheet": name, "table": table, "rows": rows,
                               "header_row": meta["header_row"],
                               "columns": columns})
                sheets_meta.append({"sheet": name, "csv": csv_path,
                                    "delimiter": meta["delimiter"],
                                    "encoding": encoding or "utf-8"})
            _write_cells(con, sheets_meta, Path(tmp))
            _write_meta(con, tables)
            return {"tables": tables}
        finally:
            con.close()


def _write_meta(con, tables: List[dict]) -> None:
    con.execute(
        f'CREATE TABLE "{META_TABLE}" ('
        ' sheet VARCHAR, table_name VARCHAR, header_row BIGINT,'
        ' row_count BIGINT, column_count BIGINT)')
    for t in tables:
        con.execute(
            f'INSERT INTO "{META_TABLE}" VALUES (?, ?, ?, ?, ?)',
            [t["sheet"], t["table"], t["header_row"], t["rows"],
             len(t["columns"])])


def a1_ref(col: int, row: int) -> str:
    """A cell's spreadsheet address — 1-based column and row to "B7"/"AA12".

    The point is that the agent and the USER can talk about the same cell. A
    person looking at the spreadsheet sees `AA12` in the corner of the window,
    not `(row=12, col=27)`; a tool that reports only the latter forces a
    translation neither side should have to do by hand. Bijective base-26, so
    26 is Z and 27 is AA — there is no "zero" letter.
    """
    letters = ""
    while col > 0:
        col, rem = divmod(col - 1, 26)
        letters = chr(ord("A") + rem) + letters
    return f"{letters}{row}"


def _safe_csv_rows(reader, meta: dict):
    """Wrap ONLY csv.reader's own `next()`. It is LAZY — an oversized field or
    an unterminated quote raises `csv.Error` while ITERATING, never at
    construction — and the caller's own per-cell work stays outside this try
    block so a bug there can never be relabelled "corrupt CSV"."""
    while True:
        try:
            row = next(reader)
        except StopIteration:
            return
        except csv.Error as e:
            raise doc_support.ParseFailed(
                f"the cell index for {meta['sheet']!r} could not be built — "
                f"a field is larger than the "
                f"{CSV_FIELD_SIZE_LIMIT // (1024 * 1024)}MB limit, or an "
                "unbalanced quote runs to the end of the file; ask the user "
                "to resend it or export a cleaner version") from e
        yield row


def _write_cells(con, sheets_meta: List[dict], workdir: Path) -> None:
    """Build the `cells` table by re-reading the very CSVs DuckDB was fed.

    Reading OUR serialization back (rather than the source a second time)
    is what makes this a real cross-check: `cells` and the typed tables are
    then two views of identical bytes, so a disagreement between them can
    only come from type inference or header detection — the two guesses this
    table exists to expose. It also keeps one code path for both formats and
    guarantees the row numbers line up with `tables_meta.header_row`.

    Sparse: only non-empty cells are stored. A 40,000-row sheet of 20 columns
    would otherwise contribute 800,000 rows recording mostly nothing, and an
    absent (row, col) already means exactly what an empty spreadsheet cell
    means.
    """
    con.execute(
        f'CREATE TABLE "{CELLS_TABLE}" ('
        ' sheet VARCHAR, row BIGINT, col BIGINT, ref VARCHAR, value VARCHAR)')
    dump = workdir / "_cells.csv"
    with open(dump, "w", newline="", encoding="utf-8") as out:
        writer = csv.writer(out, lineterminator="\n")
        for meta in sheets_meta:
            with open(meta["csv"], newline="", encoding=meta["encoding"],
                      errors="replace") as fh:
                reader = csv.reader(fh, delimiter=meta["delimiter"])
                # enumerate over LOGICAL rows: a quoted field containing a
                # newline spans several physical lines but is one row to both
                # csv.reader and DuckDB, so the numbering stays in agreement.
                for r, row in enumerate(_safe_csv_rows(reader, meta), 1):
                    for c, value in enumerate(row, 1):
                        if value != "":
                            writer.writerow(
                                [meta["sheet"], r, c, a1_ref(c, r), value])
    if dump.stat().st_size:
        con.execute(
            f'INSERT INTO "{CELLS_TABLE}" SELECT * FROM read_csv(?, '
            "header=false, columns={'sheet': 'VARCHAR', 'row': 'BIGINT', "
            "'col': 'BIGINT', 'ref': 'VARCHAR', 'value': 'VARCHAR'})",
            [str(dump)])


# How much of a result comes back. The agent asked a QUESTION; the answer to
# a good one is a handful of rows, and a query that returns hundreds is
# almost always one that should have been an aggregate instead. Truncation is
# reported, never silent, so the agent can narrow rather than assume it saw
# everything.
MAX_QUERY_ROWS = 200
QUERY_TIMEOUT_SECONDS = 15.0


class NotAQuery(Exception):
    """The submitted SQL is not a single read. Distinct from QueryFailed
    because the remedy is different in kind: this is not "your query has a
    bug", it is "this tool does not do that at all"."""
    es_code = "doc_query_not_read"


class QueryFailed(Exception):
    es_code = "doc_query_failed"


class QueryTimeout(Exception):
    es_code = "doc_query_timeout"


def _connect_readonly(db_path: Path):
    """A connection that cannot write and cannot touch the filesystem.

    Three separate things, because read-only alone is not enough. `read_only`
    stops writes to THIS database, but a plain SELECT can still reach the host
    filesystem — `SELECT * FROM read_csv('/etc/passwd')` is a read, and
    `COPY (SELECT ...) TO '/tmp/x'` is a write that never touches the
    database. `enable_external_access=false` closes both (verified: file
    reads, COPY TO and ATTACH all raise PermissionException), and DuckDB
    refuses to re-enable it at runtime — a submitted `SET
    enable_external_access=true` fails with "Cannot enable external access
    while database is running", so no lock_configuration is needed on top.
    """
    return duckdb.connect(str(db_path), read_only=True,
                          config={"enable_external_access": "false"})


def _table_summary(con) -> str:
    """The sheet -> table mapping as one line, for an error message.

    An unknown-table error is the moment the agent needs this most: it almost
    certainly typed the SHEET name it saw in the spreadsheet rather than the
    slug the table actually has, so the error has to bridge the two rather
    than just say no.
    """
    try:
        rows = con.execute(
            f'SELECT table_name, sheet FROM "{META_TABLE}" ORDER BY 1').fetchall()
    except duckdb.Error:
        return ""
    if not rows:
        return ""
    listed = ", ".join(f"{t} (sheet {s!r})" for t, s in rows)
    # Its OWN line: DuckDB's catalog error ends with a caret pointing under
    # the offending token, so appending inline puts the table list next to
    # that caret where it reads as part of the position marker rather than
    # as the answer to "then what IS it called".
    return f"\n\nTables in this document: {listed}."


def _json_safe(value):
    """MCP hands this straight to json.dumps, so a date or a Decimal has to
    already be a string or a number — failing at serialization time, after the
    query has run and the connection is closed, would report a type error in
    place of the answer."""
    if isinstance(value, (datetime.datetime, datetime.date, datetime.time)):
        return value.isoformat(sep=" ") if isinstance(
            value, datetime.datetime) else value.isoformat()
    if isinstance(value, datetime.timedelta):
        return str(value)
    if isinstance(value, (bytes, bytearray)):
        return value.decode("utf-8", "replace")
    if isinstance(value, (str, int, float, bool)) or value is None:
        return value
    return str(value)


def query(adir: Path, sql: str, *, timeout: float = QUERY_TIMEOUT_SECONDS) -> dict:
    """Run one read-only SQL statement against a document's database.

    Returns `{"columns", "rows", "row_count", "truncated"}`.

    Only `StatementType.SELECT` is accepted, as classified by DuckDB's own
    parser rather than by matching the leading word — which also means
    DESCRIBE, SHOW and SUMMARIZE come along for free (DuckDB classifies all
    three as SELECT), and those are exactly how the agent recovers a schema
    it no longer has the receipt for. Exactly ONE statement: `SELECT 1; DROP
    TABLE t` parses as two, and accepting the first would run the second.

    The row bound is applied by WRAPPING the query rather than appending a
    LIMIT — appending would attach to the wrong part of a UNION, a CTE, or a
    query that already ends in its own LIMIT. One extra row is fetched beyond
    the cap purely to tell "exactly at the cap" from "there was more".

    The time bound is a timer that calls `con.interrupt()`. DuckDB has no
    statement_timeout setting, so this was established by measurement rather
    than assumed: a three-way cross join that would never finish stopped in
    exactly the time allowed, raising InterruptException.
    """
    db_path = Path(adir) / DB_NAME
    if not db_path.is_file():
        raise QueryFailed(
            "this document has no queryable data — call es_doc_extract on "
            "the source file again to rebuild it")

    text = sql.strip().rstrip(";").strip()
    if not text:
        raise NotAQuery("no SQL was given — pass a SELECT statement")

    con = _connect_readonly(db_path)
    try:
        try:
            statements = con.extract_statements(text)
        except duckdb.Error as e:
            raise QueryFailed(f"could not parse that SQL: {e}") from e
        if len(statements) != 1:
            raise NotAQuery(
                f"send one statement at a time — that was {len(statements)}. "
                "This tool only runs a single read.")
        if statements[0].type != duckdb.StatementType.SELECT:
            raise NotAQuery(
                "this document is READ-only — only SELECT (and DESCRIBE / "
                "SHOW TABLES / SUMMARIZE) can be run against it, never "
                "INSERT, UPDATE, DELETE, DROP, CREATE, COPY or ATTACH. "
                "Nothing was changed.")

        wrapped = f"SELECT * FROM ({text}) AS es_q LIMIT {MAX_QUERY_ROWS + 1}"
        timer = threading.Timer(timeout, con.interrupt)
        timer.start()
        try:
            cur = con.execute(wrapped)
            columns = [d[0] for d in cur.description]
            fetched = cur.fetchall()
        except duckdb.InterruptException as e:
            raise QueryTimeout(
                f"the query was still running after {timeout:g}s and was "
                "stopped — narrow it (add a WHERE, or aggregate instead of "
                "listing rows)") from e
        except duckdb.Error as e:
            raise QueryFailed(f"{e}{_table_summary(con)}") from e
        finally:
            timer.cancel()
    finally:
        con.close()

    truncated = len(fetched) > MAX_QUERY_ROWS
    rows = [[_json_safe(v) for v in r] for r in fetched[:MAX_QUERY_ROWS]]
    return {"columns": columns, "rows": rows, "row_count": len(rows),
            "truncated": truncated}
